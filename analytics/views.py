from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count, F
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from sales.models import Sale, SaleItem
from inventory.models import Medicine, Batch
from purchase.models import PurchaseOrder
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


@login_required
def analytics_dashboard(request):
    today = timezone.now().date()
    last_30 = today - timedelta(days=30)
    last_7  = today - timedelta(days=7)

    # Daily revenue for chart (last 30 days)
    daily_sales = []
    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        rev = (
            Sale.objects
            .filter(sale_date__date=d, status='completed')
            .aggregate(t=Sum('total_amount'))['t'] or Decimal('0')
        )
        daily_sales.append({'date': str(d), 'revenue': float(rev)})

    # Top-selling medicines
    top_medicines = (
        SaleItem.objects
        .filter(sale__status='completed', sale__sale_date__date__gte=last_30)
        .values('medicine__name', 'medicine__brand')
        .annotate(total_qty=Sum('quantity'), total_revenue=Sum('total_price'))
        .order_by('-total_qty')[:10]
    )

    def _rev(qs_date):
        return float(
            Sale.objects
            .filter(sale_date__date__gte=qs_date, status='completed')
            .aggregate(t=Sum('total_amount'))['t'] or 0
        )

    today_rev = float(Sale.objects.filter(sale_date__date=today, status='completed')
                      .aggregate(t=Sum('total_amount'))['t'] or 0)
    week_rev  = _rev(last_7)
    month_rev = _rev(last_30)

    gst_data = (
        Sale.objects
        .filter(sale_date__date__gte=last_30, status='completed')
        .aggregate(cgst=Sum('cgst_amount'), sgst=Sum('sgst_amount'))
    )

    payment_data = list(
        Sale.objects
        .filter(sale_date__date__gte=last_30, status='completed')
        .values('payment_method')
        .annotate(count=Count('id'), total=Sum('total_amount'))
    )
    # Convert Decimal to float for JSON
    for row in payment_data:
        row['total'] = float(row['total'] or 0)

    context = {
        'daily_sales_json': json.dumps(daily_sales),
        'top_medicines': list(top_medicines),
        'today_rev': today_rev,
        'week_rev': week_rev,
        'month_rev': month_rev,
        'gst_data': {k: float(v or 0) for k, v in gst_data.items()},
        'payment_data': payment_data,
        'total_sales_count': Sale.objects.filter(
            sale_date__date__gte=last_30, status='completed').count(),
    }
    # Convert gst_data Decimal values to float for template |add filter compatibility
    context['gst_data'] = {k: round(float(v or 0), 2) for k, v in gst_data.items()}
    return render(request, 'analytics/dashboard.html', context)


@login_required
def gst_report(request):
    month = request.GET.get('month', timezone.now().strftime('%Y-%m'))
    try:
        year, mon = map(int, month.split('-'))
    except ValueError:
        year, mon = timezone.now().year, timezone.now().month

    sales = (
        Sale.objects
        .filter(sale_date__year=year, sale_date__month=mon, status='completed')
        .select_related('customer')
        .order_by('sale_date')
    )
    totals = sales.aggregate(
        subtotal=Sum('subtotal'),
        cgst=Sum('cgst_amount'),
        sgst=Sum('sgst_amount'),
        total=Sum('total_amount'),
    )
    return render(request, 'analytics/gst_report.html', {
        'sales': sales, 'totals': totals, 'month': month,
    })


@login_required
def export_gst_excel(request):
    month = request.GET.get('month', timezone.now().strftime('%Y-%m'))
    try:
        year, mon = map(int, month.split('-'))
    except ValueError:
        year, mon = timezone.now().year, timezone.now().month

    sales = (
        Sale.objects
        .filter(sale_date__year=year, sale_date__month=mon, status='completed')
        .select_related('customer')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"GST {month}"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ['Invoice No.', 'Date', 'Customer', 'Taxable Amt',
               'CGST', 'SGST', 'Total', 'Payment']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 18

    for row_idx, sale in enumerate(sales, 2):
        ws.append([
            sale.invoice_number,
            sale.sale_date.strftime('%d-%m-%Y'),
            sale.customer.name if sale.customer else 'Walk-in',
            float(sale.subtotal or 0),
            float(sale.cgst_amount or 0),
            float(sale.sgst_amount or 0),
            float(sale.total_amount or 0),
            sale.get_payment_method_display(),
        ])

    # Totals row
    total_row = sales.count() + 2
    ws.append(['', '', 'TOTAL',
               f'=SUM(D2:D{total_row-1})', f'=SUM(E2:E{total_row-1})',
               f'=SUM(F2:F{total_row-1})', f'=SUM(G2:G{total_row-1})', ''])
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="GST_{month}.xlsx"'
    wb.save(response)
    return response


@login_required
def profit_loss_report(request):
    month = request.GET.get('month', timezone.now().strftime('%Y-%m'))
    try:
        year, mon = map(int, month.split('-'))
    except ValueError:
        year, mon = timezone.now().year, timezone.now().month

    items = (
        SaleItem.objects
        .filter(sale__sale_date__year=year, sale__sale_date__month=mon,
                sale__status='completed')
        .select_related('medicine', 'batch')
    )

    report = []
    for item in items:
        cost    = float(item.batch.purchase_price) * item.quantity
        revenue = float(item.total_price)
        profit  = revenue - cost
        margin  = (profit / revenue * 100) if revenue else 0
        report.append({
            'name': item.medicine.name,
            'brand': item.medicine.brand,
            'qty': item.quantity,
            'cost': round(cost, 2),
            'revenue': round(revenue, 2),
            'profit': round(profit, 2),
            'margin': round(margin, 1),
        })

    # Sort by profit descending
    report.sort(key=lambda x: x['profit'], reverse=True)
    total_revenue = sum(r['revenue'] for r in report)
    total_cost    = sum(r['cost']    for r in report)
    total_profit  = sum(r['profit']  for r in report)

    return render(request, 'analytics/profit_loss.html', {
        'report': report,
        'total_revenue': round(total_revenue, 2),
        'total_cost': round(total_cost, 2),
        'total_profit': round(total_profit, 2),
        'month': month,
    })
